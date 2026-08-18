import json
import tempfile
import unittest
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook

from pipeline.sources.thqbca_data import parse_thqbca_workbooks


class ThqbcaParserTests(unittest.TestCase):
    def test_workbooks_become_long_records_and_exclude_compass_lookup(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            water = root / "water.xlsx"
            climate = root / "climate.xlsx"
            output = root / "observations.csv"
            manifest = root / "manifest.json"

            wb = Workbook()
            ws = wb.active
            ws.title = "TN"
            ws.append(["Date", "Whole Lake", "ML", "(mg/L)"])
            ws.append([datetime(2020, 1, 1), 1.0, 2.0, None])
            phyto = wb.create_sheet("Phyto_number")
            phyto.append(["Year", "Cyanobacteria", "Total", "(Cells/L)"])
            phyto.append([2020, 1000, 2000, None])
            wb.save(water)

            wb = Workbook()
            ws = wb.active
            ws.title = "PRE"
            ws.append(["Date", "PRE", "(mm)"])
            ws.append([datetime(2020, 1, 1), 12.0, None])
            win = wb.create_sheet("WIN")
            win.append(["Date", "Daily mean wind speed (m/s)", "Daily maximum wind speed  (m/s)", "Wind direction of daily maximum wind speed", None, "Wind direction", "Code"])
            win.append([datetime(2020, 1, 1), 3.0, None, None, None, "N", 1])
            wb.save(climate)

            result = parse_thqbca_workbooks(water, climate, output, manifest)
            rows = list(output.read_text(encoding="utf-8-sig").splitlines())
            payload = json.loads(manifest.read_text(encoding="utf-8"))

        self.assertEqual(result["records"], 5)
        self.assertIn("total_nitrogen", result["by_variable"])
        self.assertIn("algae_density", result["by_variable"])
        self.assertNotIn("wind_direction", result["by_variable"])
        self.assertIn("WIN.Wind direction", payload["excluded_fields"])
        self.assertEqual(len(rows), 6)


if __name__ == "__main__":
    unittest.main()

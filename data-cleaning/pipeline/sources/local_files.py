from __future__ import annotations

import csv
import json
import re
from datetime import date, datetime
from pathlib import Path
from typing import Any

import yaml

from ..normalize import _row
from ..time_contract import parse_time


_CONTROL_ALIASES = {
    "value", "observed_value", "clean_value", "measurement", "measure", "val", "数值",
    "variable_code", "variable", "指标", "指标编码", "监测指标", "unit", "单位",
}
_CONTROL_CANONICAL = {
    "value": "value",
    "observedvalue": "observed_value",
    "cleanvalue": "clean_value",
    "measurement": "value",
    "measure": "value",
    "val": "value",
    "数值": "value",
    "variablecode": "variable_code",
    "variable": "variable_code",
    "指标": "variable_code",
    "指标编码": "variable_code",
    "监测指标": "variable_code",
    "unit": "unit",
    "单位": "unit",
    "sourceunit": "source_unit",
    "原始单位": "source_unit",
    "conversionrule": "conversion_rule",
    "转换规则": "conversion_rule",
}
_MISSING = {"", "-", "--", "na", "n/a", "null", "none", "nan", "-999", "-999.0"}


def _key(value: Any) -> str:
    return re.sub(r"[\s_\-()/]+", "", str(value).strip()).casefold()


_CONTROL_ALIAS_KEYS = {_key(item) for item in _CONTROL_ALIASES}
_VALUE_ALIAS_KEYS = {_key(item) for item in ("value", "observed_value", "clean_value", "measurement", "measure", "val", "数值")}


def _load_alias_map(alias_path: Path | None = None) -> dict[str, str]:
    path = alias_path or Path(__file__).resolve().parents[2] / "config" / "aliases.yml"
    payload = yaml.safe_load(path.read_text(encoding="utf-8")) or {}
    result: dict[str, str] = {}
    for canonical, aliases in (payload.get("aliases") or {}).items():
        result[_key(canonical)] = canonical
        for alias in aliases or []:
            result[_key(alias)] = canonical
    return result


def _canonical(value: Any, aliases: dict[str, str]) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    key = _key(text)
    return aliases.get(key) or _CONTROL_CANONICAL.get(key) or (text if text else None)


def _read_rows(path: Path) -> list[dict[str, Any]]:
    suffix = path.suffix.casefold()
    if suffix in {".csv", ".tsv"}:
        delimiter = "\t" if suffix == ".tsv" else ","
        with path.open("r", encoding="utf-8-sig", newline="") as handle:
            return [dict(row) for row in csv.DictReader(handle, delimiter=delimiter)]
    if suffix == ".xlsx":
        from openpyxl import load_workbook

        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            sheet = workbook.active
            values = list(sheet.iter_rows(values_only=True))
            if not values:
                return []
            headers = [str(value).strip() if value is not None else "" for value in values[0]]
            return [dict(zip(headers, row)) for row in values[1:]]
        finally:
            workbook.close()
    if suffix == ".json":
        payload = json.loads(path.read_text(encoding="utf-8"))
        if isinstance(payload, list):
            return [row for row in payload if isinstance(row, dict)]
        if isinstance(payload, dict):
            for key in ("records", "rows", "data", "items"):
                if isinstance(payload.get(key), list):
                    return [row for row in payload[key] if isinstance(row, dict)]
            return [payload]
    raise ValueError(f"unsupported local file type: {path.suffix}")


def _as_time(value: Any, source_timezone: str | None = None) -> dict[str, str | None]:
    if value is None or str(value).strip() == "":
        return parse_time(value, source_timezone=source_timezone)
    if isinstance(value, (datetime, date)):
        parsed = parse_time(value, source_timezone=source_timezone)
    else:
        parsed = parse_time(str(value).strip(), source_timezone=source_timezone)
    return parsed


def _value(value: Any) -> tuple[Any, Any]:
    if value is None or str(value).strip().casefold() in _MISSING:
        return None, None
    try:
        return value, float(value)
    except (TypeError, ValueError):
        return value, value


def _coordinate(value: Any) -> Any:
    if value is None or str(value).strip() == "":
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return value


def _first(row: dict[str, Any], canonical: str, aliases: dict[str, str]) -> Any:
    for key, value in row.items():
        if _canonical(key, aliases) == canonical:
            return value
    return None


def normalize_local_file(path: Path, alias_path: Path | None = None) -> dict[str, Any]:
    """Convert a local wide or long table into the standard observation long form."""
    aliases = _load_alias_map(alias_path)
    rows = _read_rows(path)
    source_id = f"local_{path.stem}"
    observations: list[dict[str, Any]] = []
    for row_number, raw in enumerate(rows, start=2):
        # Canonicalise each header once. The previous implementation scanned
        # and regex-normalised every header again for every requested field,
        # making large canonical CSV files unnecessarily quadratic in the
        # number of metadata lookups.
        canonical_raw = {}
        for key, value in raw.items():
            canonical_key = _canonical(key, aliases)
            if canonical_key:
                canonical_raw[canonical_key] = value
        time_value = canonical_raw.get("observed_at") or canonical_raw.get("acquisition_at")
        source_timezone = canonical_raw.get("source_timezone")
        time_fields = _as_time(time_value, str(source_timezone) if source_timezone not in (None, "") else None)
        observed_at = time_fields["utc"]
        row_source = canonical_raw.get("source_id") or source_id
        station_id = canonical_raw.get("station_id")
        station_name = canonical_raw.get("station_name")
        longitude = canonical_raw.get("longitude")
        latitude = canonical_raw.get("latitude")
        unit = canonical_raw.get("unit")
        source_unit = canonical_raw.get("source_unit") or unit
        conversion_rule = canonical_raw.get("conversion_rule")
        value_origin = canonical_raw.get("value_origin") or "observed"

        variable_value = canonical_raw.get("variable_code")
        long_value = None
        for key, value in raw.items():
            key_token = _key(key)
            if key_token in _CONTROL_ALIAS_KEYS:
                if key_token in _VALUE_ALIAS_KEYS:
                    long_value = value
                    break
        if variable_value is not None and long_value is not None:
            variable_code = _canonical(variable_value, aliases)
            observed_value, clean_value = _value(long_value)
            observations.append(
                _row(
                    source_id=str(row_source), source_file=path, source_row=str(row_number),
                    observed_at=observed_at, variable_code=variable_code or "unknown_variable",
                    observed_value=observed_value, clean_value=clean_value, unit=unit,
                    value_origin=str(value_origin), station_id=str(station_id) if station_id not in (None, "") else None,
                    longitude=_coordinate(longitude),
                    latitude=_coordinate(latitude),
                    source_parameter=str(variable_value),
                    conversion_rule=conversion_rule,
                    time_fields=time_fields,
                    raw_unit=source_unit,
                )
            )
            observations[-1]["station_name"] = station_name
            observations[-1]["source_unit"] = source_unit
            observations[-1]["raw_unit"] = source_unit
            observations[-1]["conversion_rule"] = conversion_rule
            continue

        metadata = {"observed_at", "acquisition_at", "source_id", "station_id", "station_name", "lake_zone", "longitude", "latitude", "depth_m", "source_timezone", "unit", "source_unit", "conversion_rule", "value_origin", "quality_flags", "is_imputed"}
        for header, raw_value in raw.items():
            variable_code = _canonical(header, aliases)
            if not variable_code or variable_code in metadata:
                continue
            observed_value, clean_value = _value(raw_value)
            observations.append(
                _row(
                    source_id=str(row_source), source_file=path, source_row=f"{row_number}:{header}",
                    observed_at=observed_at, variable_code=variable_code,
                    observed_value=observed_value, clean_value=clean_value, unit=unit,
                    value_origin=str(value_origin), station_id=str(station_id) if station_id not in (None, "") else None,
                    longitude=_coordinate(longitude),
                    latitude=_coordinate(latitude),
                    source_parameter=header,
                    conversion_rule=conversion_rule,
                    time_fields=time_fields,
                    raw_unit=source_unit,
                )
            )
            observations[-1]["station_name"] = station_name
            observations[-1]["source_unit"] = source_unit
            observations[-1]["raw_unit"] = source_unit
    return {"observations": observations, "catalog": [], "archives": []}

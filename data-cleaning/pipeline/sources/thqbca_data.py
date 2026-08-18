from __future__ import annotations

import csv
import json
from datetime import date, datetime, timezone, timedelta
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


CN_TZ = timezone(timedelta(hours=8))
UTC = timezone.utc


WATER_SHEET_MAP = {
    "pH": ("pH", "pH"),
    "CODMn": ("cod_mn", "mg/L"),
    "DO": ("dissolved_oxygen", "mg/L"),
    "TP": ("total_phosphorus", "mg/L"),
    "PO4-P": ("phosphate_phosphorus", "mg/L"),
    "TN": ("total_nitrogen", "mg/L"),
    "NH4-N": ("ammonia_nitrogen", "mg/L"),
    "NO3-N": ("nitrate_nitrogen", "mg/L"),
    # The workbook's unit cell is mojibake for μg/L; values are in the
    # expected nitrite microgram-per-litre scale and are converted by QC.
    "NO2-N": ("nitrite_nitrogen", "ug/L"),
    "Phyto_biomass": ("phytoplankton_biomass", "mg/L"),
}


def _time(value: Any, annual: bool = False) -> str | None:
    if value is None:
        return None
    if isinstance(value, int) and annual:
        value = datetime(value, 1, 1)
    if isinstance(value, date) and not isinstance(value, datetime):
        value = datetime.combine(value, datetime.min.time())
    if not isinstance(value, datetime):
        return None
    return value.replace(tzinfo=CN_TZ).astimezone(UTC).isoformat()


def _number(value: Any) -> float | None:
    if value is None or str(value).strip().casefold() in {"", "nan", "na", "n/a", "-999", "-999.0"}:
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _row(
    *, source_file: Path, source_row: str, observed_at: str | None,
    station_id: str, variable_code: str, value: Any, unit: str,
    source_parameter: str, conversion_rule: str | None = None,
) -> dict[str, Any]:
    numeric = _number(value)
    return {
        "source_id": "taihu_thqbca_history",
        "source_file": str(source_file),
        "source_row": source_row,
        "station_id": station_id,
        "observed_at": observed_at,
        "variable_code": variable_code,
        "source_parameter": source_parameter,
        "observed_value": value,
        "clean_value": numeric,
        "unit": unit,
        "source_unit": unit,
        "conversion_rule": conversion_rule,
        "value_origin": "observed",
        "is_imputed": False,
        "imputation_method": None,
        "imputation_confidence": None,
        "quality_flags": [],
    }


def _water_quality_rows(path: Path) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        for sheet_name, (variable_code, unit) in WATER_SHEET_MAP.items():
            if sheet_name not in workbook.sheetnames:
                continue
            sheet = workbook[sheet_name]
            values = list(sheet.iter_rows(values_only=True))
            if not values:
                continue
            headers = list(values[0])
            for excel_row, data in enumerate(values[1:], start=2):
                observed_at = _time(data[0] if data else None)
                for index, header in enumerate(headers[1:], start=1):
                    if header is None or str(header).strip().startswith("("):
                        continue
                    value = data[index] if index < len(data) else None
                    station = "TAIHU_WHOLE" if str(header).strip().casefold() == "whole lake" else f"TAIHU_{str(header).strip()}"
                    rows.append(_row(
                        source_file=path, source_row=f"{sheet_name}:{excel_row}:{header}",
                        observed_at=observed_at, station_id=station,
                        variable_code=variable_code, value=value, unit=unit,
                        source_parameter=sheet_name,
                    ))

        sheet_name = "Phyto_number"
        if sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            values = list(sheet.iter_rows(values_only=True))
            headers = list(values[0]) if values else []
            cyanobacteria_index = next((i for i, header in enumerate(headers) if str(header).strip().casefold() == "cyanobacteria"), None)
            if cyanobacteria_index is not None:
                for excel_row, data in enumerate(values[1:], start=2):
                    annual_time = _time(data[0] if data else None, annual=True)
                    value = data[cyanobacteria_index] if cyanobacteria_index < len(data) else None
                    rows.append(_row(
                        source_file=path, source_row=f"{sheet_name}:{excel_row}:Cyanobacteria",
                        observed_at=annual_time, station_id="TAIHU_WHOLE",
                        variable_code="algae_density", value=value, unit="cells/L",
                        source_parameter="Cyanobacteria",
                        conversion_rule="annual aggregate; date anchored to Jan 1 Asia/Shanghai",
                    ))
    finally:
        workbook.close()
    return rows


def _climate_rows(path: Path) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        simple = {
            "PRE": ("precipitation", "mm", 1, "PRE"),
            "TEM": ("air_temperature", "degC", 1, "TEM"),
            "WaterLevel": ("water_level", "m", 1, "Water Depth"),
        }
        for sheet_name, (variable_code, unit, value_index, source_parameter) in simple.items():
            if sheet_name not in workbook.sheetnames:
                continue
            sheet = workbook[sheet_name]
            for excel_row, data in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                value = data[value_index] if value_index < len(data) else None
                rule = "source label Water Depth; datum semantics require validation" if sheet_name == "WaterLevel" else None
                rows.append(_row(
                    source_file=path, source_row=f"{sheet_name}:{excel_row}",
                    observed_at=_time(data[0] if data else None), station_id="TAIHU_CLIMATE" if sheet_name != "WaterLevel" else "TAIHU_WATER_LEVEL",
                    variable_code=variable_code, value=value, unit=unit,
                    source_parameter=source_parameter, conversion_rule=rule,
                ))

        if "WIN" in workbook.sheetnames:
            sheet = workbook["WIN"]
            for excel_row, data in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                observed_at = _time(data[0] if data else None)
                speed = data[1] if len(data) > 1 else None
                rows.append(_row(
                    source_file=path, source_row=f"WIN:{excel_row}:mean_speed",
                    observed_at=observed_at, station_id="TAIHU_CLIMATE",
                    variable_code="wind_speed", value=speed, unit="m/s",
                    source_parameter="Daily mean wind speed (m/s)",
                ))
    finally:
        workbook.close()
    return rows


def parse_thqbca_workbooks(water_quality: Path, climate: Path, output_csv: Path, manifest_path: Path) -> dict[str, Any]:
    records = _water_quality_rows(water_quality) + _climate_rows(climate)
    fields = [
        "source_id", "source_file", "source_row", "station_id", "observed_at",
        "variable_code", "source_parameter", "observed_value", "clean_value",
        "unit", "source_unit", "conversion_rule", "value_origin", "is_imputed",
        "imputation_method", "imputation_confidence", "quality_flags",
    ]
    output_csv.parent.mkdir(parents=True, exist_ok=True)
    with output_csv.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fields)
        writer.writeheader()
        for row in records:
            writer.writerow({key: json.dumps(row[key], ensure_ascii=False) if isinstance(row[key], (list, dict)) else row[key] for key in fields})
    by_variable: dict[str, int] = {}
    missing_by_variable: dict[str, int] = {}
    for row in records:
        code = row["variable_code"]
        by_variable[code] = by_variable.get(code, 0) + 1
        if row["clean_value"] is None:
            missing_by_variable[code] = missing_by_variable.get(code, 0) + 1
    payload = {
        "source_id": "taihu_thqbca_history",
        "water_quality_file": str(water_quality),
        "climate_file": str(climate),
        "output_csv": str(output_csv),
        "records": len(records),
        "by_variable": by_variable,
        "missing_by_variable": missing_by_variable,
        "missing_rate_by_variable": {key: missing_by_variable.get(key, 0) / count for key, count in by_variable.items()},
        "time_semantics": "water quality monthly/quarterly; climate daily; Phyto_number annual anchored to Jan 1",
        "excluded_fields": {
            "WIN.Wind direction": "workbook contains only a 16-row compass lookup at the beginning, followed by nulls; not treated as daily observations",
        },
    }
    manifest_path.parent.mkdir(parents=True, exist_ok=True)
    manifest_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    return payload

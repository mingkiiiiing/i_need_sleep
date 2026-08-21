from __future__ import annotations

import json
import sqlite3
from datetime import datetime
from pathlib import Path
from typing import Any
from zoneinfo import ZoneInfo

import pandas as pd
from openpyxl import load_workbook

UTC = ZoneInfo("UTC")
LOCAL = ZoneInfo("Asia/Shanghai")
WAVELENGTHS = {"rrs_blue_490": 490, "rrs_green_560": 560, "rrs_red_665": 665, "rrs_rededge_705": 705, "rrs_nir_842": 842}


def _time(value: Any) -> str | None:
    text = str(value or "").strip().replace("_", "-")
    for pattern in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M"):
        try:
            return datetime.strptime(" ".join(text.split()), pattern).replace(tzinfo=LOCAL).astimezone(UTC).isoformat()
        except ValueError:
            continue
    return None


def _number(value: Any) -> float | None:
    try:
        result = float(value)
        return result if pd.notna(result) else None
    except (TypeError, ValueError):
        return None


def parse_workbook(path: Path) -> list[dict[str, Any]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        quality = list(workbook["Water quality dataset"].iter_rows(values_only=True))
        spectra = list(workbook["Water spectral dataset"].iter_rows(values_only=True))
    finally:
        workbook.close()
    q_header = [str(item or "") for item in quality[0]]
    s_header = list(spectra[0])
    wavelength_index = {}
    for name, wavelength in WAVELENGTHS.items():
        wavelength_index[name] = min(range(4, len(s_header)), key=lambda index: abs(float(s_header[index]) - wavelength) if str(s_header[index]).replace(".", "", 1).isdigit() else 1e9)
    rows = []
    for position, qrow in enumerate(quality[1:]):
        if position >= len(spectra) - 1:
            break
        srow = spectra[position + 1]
        observed_at = _time(qrow[1] if len(qrow) > 1 else None)
        row = {
            "sample_id": f"{path.stem}_{position:03d}", "source_id": "zenodo_taihu_satellite_ground_insitu",
            "source_file": str(path), "source_row": position + 2, "observed_at": observed_at,
            "longitude": _number(qrow[2]), "latitude": _number(qrow[3]),
            "chlorophyll_a_ug_l": _number(qrow[4]), "total_suspended_matter_mg_l": _number(qrow[5]),
            "secchi_depth_cm": _number(qrow[6]), "water_temperature_c": _number(qrow[7]),
        }
        for name, index in wavelength_index.items():
            row[name] = _number(srow[index]) if index < len(srow) else None
        red, rededge, nir = row["rrs_red_665"], row["rrs_rededge_705"], row["rrs_nir_842"]
        row["ndci_field"] = (rededge - red) / (rededge + red) if red is not None and rededge is not None and rededge + red != 0 else None
        row["mci_field"] = rededge - red - (nir - red) * ((705 - 665) / (842 - 665)) if None not in (red, rededge, nir) else None
        rows.append(row)
    return rows


def run_taihu_insitu_parse(input_root: Path, output_root: Path, database: Path, *, manifest_path: Path | None = None) -> dict[str, Any]:
    paths = sorted(input_root.glob("Lake Taihu*_Water_dataset.xlsx"))
    rows = [row for path in paths for row in parse_workbook(path)]
    frame = pd.DataFrame(rows)
    output_root.mkdir(parents=True, exist_ok=True)
    csv_path, parquet_path = output_root / "taihu_insitu_calibration_samples.csv", output_root / "taihu_insitu_calibration_samples.parquet"
    frame.to_csv(csv_path, index=False, encoding="utf-8-sig")
    frame.to_parquet(parquet_path, index=False)
    database.parent.mkdir(parents=True, exist_ok=True)
    with sqlite3.connect(database) as connection:
        frame.to_sql("taihu_insitu_calibration_samples", connection, if_exists="replace", index=False)
        connection.execute("CREATE UNIQUE INDEX IF NOT EXISTS idx_taihu_insitu_sample ON taihu_insitu_calibration_samples(sample_id)")
    manifest_path = manifest_path or output_root / "manifest.json"
    manifest = {"source_id": "zenodo_taihu_satellite_ground_insitu", "status": "completed" if rows else "BLOCKED_DATA", "input_files": [str(path) for path in paths], "rows": len(rows), "chlorophyll_rows": int(frame["chlorophyll_a_ug_l"].notna().sum()) if len(frame) else 0, "date_count": int(frame["observed_at"].str[:10].nunique()) if len(frame) else 0, "outputs": {"csv": str(csv_path), "parquet": str(parquet_path), "database": str(database)}, "manifest": str(manifest_path)}
    manifest_path.write_text(json.dumps(manifest, ensure_ascii=False, indent=2), encoding="utf-8")
    return manifest

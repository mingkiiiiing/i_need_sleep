from __future__ import annotations

"""Read-only file-level quarantine checks for the Taihu raw-data boundary.

The module deliberately does not move or delete files.  It inventories every
file under a bounded input root, computes a SHA-256 checksum, and records
issues that must be resolved before tabular cleaning (empty files, unreadable
encoding, malformed schema, corrupt archives, and duplicate payloads).
"""

import csv
import gzip
import hashlib
import json
import shutil
import zipfile
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Iterable


STORAGE = Path(__import__("os").environ.get("TAIHU_STORAGE_ROOT") or (Path(__file__).resolve().parents[1] / "storage"))
DEFAULT_INPUT_ROOT = STORAGE / "raw"
DEFAULT_REPORT = STORAGE / "reports" / "file_quarantine.csv"
DEFAULT_MANIFEST = STORAGE / "manifests" / "file_quarantine.json"

TEXT_EXTENSIONS = {".csv", ".tsv", ".json", ".html", ".htm", ".md", ".txt", ".xml", ".yml", ".yaml", ".cpg", ".prj"}
TABULAR_EXTENSIONS = {".csv", ".tsv"}
JSON_EXTENSIONS = {".json"}
EXCEL_EXTENSIONS = {".xlsx", ".xlsm"}
ZIP_EXTENSIONS = {".zip"}
GZIP_EXTENSIONS = {".gz"}
RAR_EXTENSIONS = {".rar"}

REPORT_FIELDS = [
    "relative_path",
    "absolute_path",
    "suffix",
    "size_bytes",
    "sha256",
    "status",
    "issue_codes",
    "encoding",
    "schema_status",
    "row_count",
    "duplicate_of",
    "archive_check",
    "checked_at_utc",
    "error",
]


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _decode_text(path: Path) -> tuple[str | None, str | None]:
    data = path.read_bytes()
    for encoding in ("utf-8-sig", "utf-8", "gb18030"):
        try:
            data.decode(encoding, errors="strict")
            return encoding, None
        except UnicodeDecodeError:
            continue
    return None, "text cannot be decoded as UTF-8/GB18030"


def _check_tabular(path: Path, encoding: str | None) -> tuple[str, int | None, list[str], str | None]:
    if not encoding:
        return "unreadable", None, ["encoding_unreadable", "schema_unreadable"], None
    delimiter = "\t" if path.suffix.lower() == ".tsv" else ","
    issues: list[str] = []
    rows = 0
    try:
        with path.open("r", encoding=encoding, newline="", errors="strict") as handle:
            reader = csv.reader(handle, delimiter=delimiter, strict=True)
            try:
                header = next(reader)
            except StopIteration:
                return "empty", 0, ["schema_empty"], None
            normalized = [item.strip() for item in header]
            if not normalized or all(not item for item in normalized):
                issues.append("schema_missing_header")
            if len(set(normalized)) != len(normalized):
                issues.append("schema_duplicate_columns")
            width = len(header)
            for row in reader:
                rows += 1
                if len(row) != width:
                    issues.append("schema_row_width_mismatch")
                    break
            if rows == 0:
                issues.append("schema_no_rows")
        return ("invalid" if issues else "valid"), rows, sorted(set(issues)), None
    except (csv.Error, UnicodeError, OSError) as exc:
        return "unreadable", rows, ["schema_unreadable"], str(exc)


def _check_json(path: Path, encoding: str | None) -> tuple[str, int | None, list[str], str | None]:
    if not encoding:
        return "unreadable", None, ["encoding_unreadable", "schema_unreadable"], None
    try:
        with path.open("r", encoding=encoding, errors="strict") as handle:
            payload = json.load(handle)
        if isinstance(payload, list):
            rows = len(payload)
            issues = ["schema_no_rows"] if rows == 0 else []
            return ("empty" if rows == 0 else "valid_json"), rows, issues, None
        if isinstance(payload, dict):
            return "valid_json", 1, [], None
        return "invalid", None, ["schema_json_scalar"], None
    except (json.JSONDecodeError, UnicodeError, OSError) as exc:
        return "unreadable", None, ["schema_unreadable"], str(exc)


def _check_excel(path: Path) -> tuple[str, int | None, list[str], str | None]:
    try:
        from openpyxl import load_workbook

        workbook = load_workbook(path, read_only=True, data_only=True)
        rows = 0
        sheets = 0
        for sheet in workbook.worksheets:
            sheets += 1
            rows += max(0, int(sheet.max_row or 0))
        workbook.close()
        if sheets == 0:
            return "invalid", rows, ["schema_no_sheets"], None
        if rows == 0:
            return "empty", rows, ["schema_no_rows"], None
        return "valid_workbook", rows, [], None
    except Exception as exc:  # openpyxl emits several format-specific errors
        return "unreadable", None, ["schema_unreadable", "compressed_or_container_corrupt"], str(exc)


def _check_archive(path: Path) -> tuple[str, list[str], str | None]:
    suffix = path.suffix.lower()
    if suffix in ZIP_EXTENSIONS:
        try:
            with zipfile.ZipFile(path) as archive:
                bad_member = archive.testzip()
                if bad_member:
                    return "corrupt", ["compressed_corrupt"], f"first_bad_member={bad_member}"
                if not archive.namelist():
                    return "valid_empty", ["archive_empty"], None
            return "valid", [], None
        except (zipfile.BadZipFile, OSError) as exc:
            return "corrupt", ["compressed_corrupt"], str(exc)
    if suffix in GZIP_EXTENSIONS:
        try:
            with gzip.open(path, "rb") as handle:
                while handle.read(1024 * 1024):
                    pass
            return "valid", [], None
        except (OSError, EOFError) as exc:
            return "corrupt", ["compressed_corrupt"], str(exc)
    if suffix in RAR_EXTENSIONS:
        try:
            signature = path.open("rb").read(7)
        except OSError as exc:
            return "unreadable", ["archive_unreadable"], str(exc)
        if signature.startswith(b"Rar!"):
            # RAR verification requires an external extractor.  Do not call an
            # untrusted binary in the pipeline; leave an explicit audit trail.
            return "not_checked", ["archive_check_unsupported"], "RAR integrity check requires authorized extractor"
        return "corrupt", ["compressed_corrupt"], "RAR signature missing"
    return "not_applicable", [], None


def _iter_files(root: Path) -> Iterable[Path]:
    if root.is_file():
        yield root
        return
    for path in sorted(root.rglob("*")):
        if path.is_file():
            yield path


def _check_file(path: Path, root: Path, seen: dict[str, str], checked_at: str) -> dict[str, Any]:
    relative = path.relative_to(root).as_posix() if path != root else path.name
    suffix = path.suffix.lower()
    issues: list[str] = []
    encoding: str | None = None
    schema_status = "not_applicable"
    row_count: int | None = None
    archive_check = "not_applicable"
    error: str | None = None
    size = path.stat().st_size
    checksum: str | None = None
    status = "ok"
    try:
        checksum = _sha256(path)
    except OSError as exc:
        issues.append("checksum_unreadable")
        error = str(exc)
    if size == 0:
        issues.append("empty_file")
        schema_status = "empty"
    elif suffix in ZIP_EXTENSIONS | GZIP_EXTENSIONS | RAR_EXTENSIONS:
        archive_check, archive_issues, archive_error = _check_archive(path)
        issues.extend(archive_issues)
        error = error or archive_error
    elif suffix in TEXT_EXTENSIONS:
        encoding, decode_error = _decode_text(path)
        if decode_error:
            issues.append("encoding_unreadable")
            error = error or decode_error
        elif encoding not in {"utf-8", "utf-8-sig"}:
            issues.append("encoding_non_utf8")
        if suffix in TABULAR_EXTENSIONS:
            schema_status, row_count, schema_issues, schema_error = _check_tabular(path, encoding)
            issues.extend(schema_issues)
            error = error or schema_error
        elif suffix in JSON_EXTENSIONS:
            schema_status, row_count, schema_issues, schema_error = _check_json(path, encoding)
            issues.extend(schema_issues)
            error = error or schema_error
    elif suffix in EXCEL_EXTENSIONS:
        schema_status, row_count, schema_issues, schema_error = _check_excel(path)
        issues.extend(schema_issues)
        error = error or schema_error

    duplicate_of = None
    if checksum:
        duplicate_of = seen.get(checksum)
        if duplicate_of:
            issues.append("duplicate_checksum")
        else:
            seen[checksum] = relative
    if issues:
        status = "duplicate" if issues == ["duplicate_checksum"] else "issue"
    return {
        "relative_path": relative,
        "absolute_path": str(path.resolve()),
        "suffix": suffix,
        "size_bytes": size,
        "sha256": checksum or "",
        "status": status,
        "issue_codes": ";".join(sorted(set(issues))),
        "encoding": encoding or "",
        "schema_status": schema_status,
        "row_count": "" if row_count is None else row_count,
        "duplicate_of": duplicate_of or "",
        "archive_check": archive_check,
        "checked_at_utc": checked_at,
        "error": error or "",
    }


def run_file_quarantine(
    input_root: Path = DEFAULT_INPUT_ROOT,
    report_path: Path = DEFAULT_REPORT,
    manifest_path: Path = DEFAULT_MANIFEST,
) -> dict[str, Any]:
    """Run read-only checks and write CSV + JSON audit artifacts."""

    root = Path(input_root)
    report_path = Path(report_path)
    manifest_path = Path(manifest_path)
    checked_at = datetime.now(timezone.utc).replace(microsecond=0).isoformat().replace("+00:00", "Z")
    if not root.exists():
        result = {"status": "BLOCKED_DATA", "input_root": str(root), "file_count": 0, "error": "input root does not exist"}
        manifest_path.parent.mkdir(parents=True, exist_ok=True)
        manifest_path.write_text(json.dumps(result, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
        return result
    seen: dict[str, str] = {}
    rows = [_check_file(path, root, seen, checked_at) for path in _iter_files(root)]
    report_path.parent.mkdir(parents=True, exist_ok=True)
    with report_path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=REPORT_FIELDS)
        writer.writeheader()
        writer.writerows(rows)
    issue_count = sum(1 for row in rows if row["status"] in {"issue", "duplicate"})
    duplicate_count = sum(1 for row in rows if "duplicate_checksum" in row["issue_codes"])
    manifest = {
        "schema_version": "1.0",
        "manifest_type": "file_quarantine",
        "status": "completed_with_issues" if issue_count else "completed",
        "input_root": str(root.resolve()),
        "report_path": str(report_path),
        "checked_at_utc": checked_at,
        "file_count": len(rows),
        "issue_count": issue_count,
        "duplicate_count": duplicate_count,
        "issue_code_counts": _issue_counts(rows),
        "read_only": True,
    }
    manifest_path.parent.mkdir(parents=True, exist_ok=True)
    manifest_path.write_text(json.dumps(manifest, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    return manifest


def _issue_counts(rows: list[dict[str, Any]]) -> dict[str, int]:
    counts: dict[str, int] = {}
    for row in rows:
        for code in str(row.get("issue_codes", "")).split(";"):
            if code:
                counts[code] = counts.get(code, 0) + 1
    return dict(sorted(counts.items()))


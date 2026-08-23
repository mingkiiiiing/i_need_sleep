# -*- coding: utf-8 -*-
"""水文数据清洗。

- THQBCA-V2 3.Climate.xlsx「WaterLevel」表: 太湖逐日平均水位(m), 2004-2020
- 水利部水情接口批次(mwr_hfc, GBK CSV): 湖面站 TH-01 等 水位/流量 实时值 (2026-08)
- tba_hydrology 目录内为下载失败响应(406/403 HTML), 无有效数据, 记录到质量报告。

输出: storage/cleaned/hydrology_cleaned.csv (+ .parquet)
"""
from __future__ import annotations

import sys
from pathlib import Path

import numpy as np
import pandas as pd

sys.path.insert(0, str(Path(__file__).resolve().parent))
from lib_common import (
    ROOT, STORAGE, bnow, coerce_datetime, date_of, datetime_of, flag_join,
    month_of, read_table, to_number, write_dataset,
)

SOURCE_WL = "taihu_thqbca_v2_water_level"
SOURCE_MWR = "mwr_hfc_water_station"


def clean_thqbca_waterlevel(rows: list[dict]) -> None:
    xlsx = list((ROOT / "storage/THQBCA-V2/3.Climate").glob("3.Climate*.xlsx"))
    if not xlsx:
        xlsx = list((ROOT / "storage/raw/taihu_thqbca_zenodo/extracted/THQBCA-V2/3.Climate").glob("3.Climate*.xlsx"))
    if not xlsx:
        print("  [水文] 未找到 3.Climate.xlsx")
        return
    import openpyxl
    wb = openpyxl.load_workbook(xlsx[0], data_only=True, read_only=True)
    ws = wb["WaterLevel"]
    n = 0
    for r in ws.iter_rows(values_only=True):
        if r[0] is None or r[1] is None:
            continue
        ts = coerce_datetime(r[0])
        val = to_number(r[1])
        if pd.isna(ts) or np.isnan(val):
            continue
        flags = flag_join(["Q00"])
        rows.append(dict(
            station_id="TAIHU_WATER_LEVEL", station_name="太湖平均水位站",
            observed_at=datetime_of(ts), date=date_of(ts), month=month_of(ts),
            variable_code="water_level", value=val, value_text="",
            unit="m", quality_flag=flags, quality_note="THQBCA数据集湖北区平均水位",
            source_name=SOURCE_WL, source_file=str(xlsx[0].relative_to(ROOT)),
            source_row=str(ws.max_row or 0), source_unit="m",
            conversion_rule="", value_origin="observed",
            longitude=float("nan"), latitude=float("nan"),
            acquisition_date=bnow().strftime("%Y-%m-%d %H:%M:%S"),
        ))
        n += 1
    wb.close()
    print(f"  [THQBCA水位] {n} 行")


def clean_mwr(rows: list[dict]) -> None:
    files = sorted((STORAGE / "raw/mwr_hfc").glob("*.csv"))
    cnt = 0
    for p in files:
        try:
            df, enc = read_table(p)
        except Exception as e:
            print(f"  [mwr] {p.name} 读取失败: {e}")
            continue
        if df.empty or df.shape[1] < 5:
            continue
        cols = {str(c).strip(): i for i, c in enumerate(df.columns)}
        # 中文表头(GBK): 站编号,站名,时间,指标,数值,单位,状态
        try:
            for _, r in df.iterrows():
                vals = [str(v) for v in r.tolist()]
                sid = vals[0].strip()
                sname = vals[1].strip() if len(vals) > 1 else ""
                tstr = vals[2].strip() if len(vals) > 2 else ""
                indicator = vals[3].strip() if len(vals) > 3 else ""
                raw_val = vals[4].strip() if len(vals) > 4 else ""
                unit = vals[5].strip() if len(vals) > 5 else ""
                ts = coerce_datetime(tstr)
                val = to_number(raw_val)
                if pd.isna(ts) or np.isnan(val):
                    continue
                indicator_std = {"水位": "water_level", "流量": "flow_rate"}.get(indicator, indicator)
                cnt += 1
                rows.append(dict(
                    station_id=sid, station_name=sname,
                    observed_at=datetime_of(ts), date=date_of(ts), month=month_of(ts),
                    variable_code=indicator_std, value=val, value_text="",
                    unit=unit or "", quality_flag="Q00",
                    quality_note="水利部水情接口实时批次", source_name=SOURCE_MWR,
                    source_file=str(p.relative_to(ROOT)), source_row=str(int(_)) if False else str(_),
                    source_unit=unit or "", conversion_rule="", value_origin="observed",
                    longitude=float("nan"), latitude=float("nan"),
                    acquisition_date=bnow().strftime("%Y-%m-%d %H:%M:%S"),
                ))
        except Exception as e:
            print(f"  [mwr] {p.name} 行解析失败: {e}")
            continue
    print(f"  [mwr洪] {cnt} 行")


def main() -> pd.DataFrame:
    print("== 水文清洗 ==")
    rows: list[dict] = []
    clean_thqbca_waterlevel(rows)
    clean_mwr(rows)
    df = pd.DataFrame(rows)
    if df.empty:
        df = pd.DataFrame(columns=["station_id"])
    df["_t"] = pd.to_datetime(df["observed_at"], errors="coerce")
    df = df.sort_values(["_t", "station_id", "variable_code"]).drop(columns="_t").reset_index(drop=True)
    path = write_dataset(df, "hydrology_cleaned")
    print(f"  [输出] {path}  {df.shape[0]} 行 x {df.shape[1]} 列;"
          f" 时间 {df['observed_at'].min()} .. {df['observed_at'].max()}" if len(df) else "  [输出] 无数据")
    return df


if __name__ == "__main__":
    main()
